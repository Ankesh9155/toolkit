"""
Core processing logic for the Suppression & TAL Toolkit.
All functions take raw bytes + filename in, return pandas DataFrames / dicts out.
No file-system dependency beyond what the caller provides.
"""

import hashlib
import io
import re
import zipfile
from collections import defaultdict

import openpyxl
import pandas as pd

MD5_RE = re.compile(r"^[a-fA-F0-9]{32}$")
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
DOMAIN_RE = re.compile(r"^(?:[a-z0-9-]+\.)+[a-z]{2,}$", re.IGNORECASE)

# Uploads larger than this are parsed in row batches instead of being handed
# to pandas whole, so peak memory during parsing stays bounded regardless of
# how big the source file is (matters on Render's 512MB instance).
CHUNK_SIZE_BYTES = 9 * 1024 * 1024  # 9 MB
CHUNK_ROWS = 20_000


# ---------------------------------------------------------------------------
# File reading helpers
# ---------------------------------------------------------------------------

def _read_xlsx_chunked(content: bytes) -> dict:
    """Stream an .xlsx/.xlsm workbook using openpyxl's read-only row iterator,
    building each sheet's DataFrame out of row batches instead of letting
    pandas buffer the whole sheet in memory before constructing it. Only used
    for uploads over CHUNK_SIZE_BYTES; smaller files go through the normal
    pandas path since the overhead isn't worth it there."""
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        tables = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            row_iter = ws.iter_rows(values_only=True)
            try:
                header = next(row_iter)
            except StopIteration:
                tables[sheet_name] = pd.DataFrame()
                continue
            columns = [c if c is not None else f"Unnamed: {i}" for i, c in enumerate(header)]

            chunks = []
            batch = []
            for row in row_iter:
                batch.append(row)
                if len(batch) >= CHUNK_ROWS:
                    chunks.append(pd.DataFrame(batch, columns=columns))
                    batch = []
            if batch:
                chunks.append(pd.DataFrame(batch, columns=columns))

            tables[sheet_name] = pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame(columns=columns)
        return tables
    finally:
        wb.close()


def _read_csv_chunked(content: bytes) -> pd.DataFrame:
    """Same idea as _read_xlsx_chunked but for CSV, via pandas' native
    chunksize reader instead of loading the whole file in one read_csv call."""
    reader = pd.read_csv(io.BytesIO(content), chunksize=CHUNK_ROWS)
    return pd.concat(reader, ignore_index=True)


def _header_looks_like_data(df: pd.DataFrame) -> bool:
    """Detect the common 'headerless file' trap: pandas treated the first
    real value (an email or an MD5 hash) as the column name, because the
    file has no header row at all."""
    for col in df.columns:
        kind_norm = classify(str(col))
        if kind_norm and kind_norm[0] in ("email", "md5"):
            return True
    return False


def _promote_header_to_row(df: pd.DataFrame) -> pd.DataFrame:
    """Undo the header-consumed-a-data-row mistake: push the column names
    back in as the first row, and rename columns generically."""
    generic_cols = [f"col_{i}" for i in range(len(df.columns))]
    header_row = pd.DataFrame([list(df.columns)], columns=generic_cols)
    df = df.copy()
    df.columns = generic_cols
    return pd.concat([header_row, df], ignore_index=True)


def read_tables(content: bytes, filename: str) -> dict:
    """Return {sheet_name: DataFrame}. CSV/TXT -> one 'Sheet1' entry.
    XLSX -> one entry per sheet (multi-tab support).
    Auto-recovers files that have no header row (single column of raw
    emails/hashes) which pandas would otherwise misread as a header."""
    name = filename.lower()
    if name.endswith((".xlsx", ".xlsm")) and len(content) > CHUNK_SIZE_BYTES:
        tables = _read_xlsx_chunked(content)
    elif name.endswith((".xlsx", ".xlsm", ".xls")):
        engine_kwargs = {"read_only": True} if name.endswith((".xlsx", ".xlsm")) else None
        xls = pd.ExcelFile(io.BytesIO(content), engine_kwargs=engine_kwargs)
        tables = {sheet: xls.parse(sheet) for sheet in xls.sheet_names}
    elif name.endswith(".csv"):
        if len(content) > CHUNK_SIZE_BYTES:
            tables = {"Sheet1": _read_csv_chunked(content)}
        else:
            tables = {"Sheet1": pd.read_csv(io.BytesIO(content))}
    elif name.endswith(".txt"):
        lines = [l.strip() for l in content.decode("utf-8", errors="ignore").splitlines() if l.strip()]
        tables = {"Sheet1": pd.DataFrame({"value": lines})}
    else:
        raise ValueError(f"Unsupported file type: {filename}")

    for sheet_name, df in list(tables.items()):
        if _header_looks_like_data(df):
            tables[sheet_name] = _promote_header_to_row(df)

    return tables


def read_single_table(content: bytes, filename: str) -> pd.DataFrame:
    """For files where we only care about the first/only sheet."""
    tables = read_tables(content, filename)
    first_key = list(tables.keys())[0]
    return tables[first_key]


# ---------------------------------------------------------------------------
# Value classification
# ---------------------------------------------------------------------------

def classify(value: str):
    v = str(value).strip()
    if not v or v.lower() == "nan":
        return None
    if MD5_RE.match(v):
        return ("md5", v.lower())
    if EMAIL_RE.match(v):
        return ("email", v.lower())
    return ("unknown", v)


def md5_of(text: str) -> str:
    return hashlib.md5(text.strip().lower().encode("utf-8")).hexdigest()


def domain_of_email(email: str) -> str:
    return email.split("@")[-1].strip().lower() if "@" in email else ""


def domain_of_url(url: str) -> str:
    u = str(url).strip().lower()
    u = re.sub(r"^https?://", "", u)
    u = re.sub(r"^www\.", "", u)
    u = u.split("/")[0]
    return u


# ---------------------------------------------------------------------------
# Column auto-detection
# ---------------------------------------------------------------------------

def detect_email_column(df: pd.DataFrame) -> str:
    best_col, best_score = None, -1
    for col in df.columns:
        if "email" in str(col).lower():
            return col
    for col in df.columns:
        sample = df[col].dropna().astype(str).head(50)
        if len(sample) == 0:
            continue
        matches = sum(1 for v in sample if EMAIL_RE.match(v.strip()))
        score = matches / len(sample)
        if score > best_score:
            best_score, best_col = score, col
    return best_col if best_score and best_score > 0.3 else (df.columns[0] if len(df.columns) else None)


def detect_domain_column(df: pd.DataFrame):
    """Look for an explicit domain/website column, else an email column to derive from."""
    name_priority = ["domain", "website", "url", "company domain"]
    for pat in name_priority:
        for col in df.columns:
            if pat in str(col).lower():
                return ("domain", col)
    for col in df.columns:
        if "email" in str(col).lower():
            return ("email", col)
    # fallback: sample values that look like domains
    for col in df.columns:
        sample = df[col].dropna().astype(str).head(50)
        if len(sample) == 0:
            continue
        matches = sum(1 for v in sample if DOMAIN_RE.match(domain_of_url(v)))
        if matches / len(sample) > 0.3:
            return ("domain", col)
    return (None, None)


# ---------------------------------------------------------------------------
# Tool 1: Email Scrubber (suppression vs lead list)
# ---------------------------------------------------------------------------

def build_suppression_sets(content: bytes, filename: str):
    """Suppression file can be CSV/TXT (one value per line) or XLSX (auto-detect column).
    Values can be plain emails or MD5 hashes, mixed."""
    tables = read_tables(content, filename)
    supp_emails, supp_hashes = set(), set()
    for df in tables.values():
        if df.shape[1] == 1:
            values = df.iloc[:, 0].dropna().astype(str).tolist()
        else:
            col = detect_email_column(df)
            values = df[col].dropna().astype(str).tolist() if col else []
        for v in values:
            kind_norm = classify(v)
            if not kind_norm:
                continue
            kind, norm = kind_norm
            if kind == "email":
                supp_emails.add(norm)
            elif kind == "md5":
                supp_hashes.add(norm)
    return supp_emails, supp_hashes


def find_duplicate_reasons(df: pd.DataFrame, col: str) -> dict:
    """Return {row_index: reason_string} for rows that are duplicates WITHIN
    a single list's email column: exact repeats (same email or same MD5
    twice) and cross-format repeats (a plain email and its own MD5 hash both
    present elsewhere in the same list)."""
    email_to_rows = defaultdict(list)
    hash_to_rows = defaultdict(list)

    for idx, raw in df[col].astype(str).items():
        kind_norm = classify(raw)
        if not kind_norm:
            continue
        kind, norm = kind_norm
        if kind == "email":
            email_to_rows[norm].append(idx)
        elif kind == "md5":
            hash_to_rows[norm].append(idx)

    reasons = defaultdict(list)

    for norm, rows in email_to_rows.items():
        if len(rows) > 1:
            for r in rows:
                reasons[r].append("duplicate email")
    for norm, rows in hash_to_rows.items():
        if len(rows) > 1:
            for r in rows:
                reasons[r].append("duplicate hash")
    for norm, rows in email_to_rows.items():
        h = md5_of(norm)
        if h in hash_to_rows:
            for r in rows:
                reasons[r].append("matches an MD5 hash elsewhere in this list")
            for r in hash_to_rows[h]:
                reasons[r].append("is the MD5 hash of an email elsewhere in this list")

    return {idx: ", ".join(dict.fromkeys(vals)) for idx, vals in reasons.items()}


def find_duplicates_in_column(df: pd.DataFrame, col: str) -> pd.DataFrame:
    """Return just the duplicate rows (with a duplicate_reason column)."""
    reasons = find_duplicate_reasons(df, col)
    if not reasons:
        return df.iloc[0:0].copy()
    dup_idx = sorted(reasons.keys())
    dup_df = df.loc[dup_idx].copy()
    dup_df["duplicate_reason"] = [reasons[i] for i in dup_idx]
    return dup_df


def email_scrubber(supp_content, supp_name, lead_content, lead_name, email_column=None):
    supp_emails, supp_hashes = build_suppression_sets(supp_content, supp_name)
    leads_df = read_single_table(lead_content, lead_name)

    col = email_column if email_column and email_column in leads_df.columns else detect_email_column(leads_df)
    if not col:
        raise ValueError("Could not detect an email column in the lead list.")

    def is_suppressed(v):
        kind_norm = classify(v)
        if not kind_norm:
            return False
        kind, norm = kind_norm
        if kind != "email":
            return False
        return norm in supp_emails or md5_of(norm) in supp_hashes

    mask = leads_df[col].astype(str).apply(is_suppressed)

    dup_reasons = find_duplicate_reasons(leads_df, col)
    duplicates_df = find_duplicates_in_column(leads_df, col)

    # Full annotated view: every lead, with its suppression + duplicate status,
    # for displaying the whole list in the UI (not just the duplicates).
    # leads_df isn't needed unmodified after this point, so annotate it in
    # place instead of holding a second full copy alongside it.
    leads_df["Found in suppression"] = mask.map({True: "Yes", False: "No"})
    leads_df["Duplicate"] = leads_df.index.map(lambda i: "Yes" if i in dup_reasons else "No")
    leads_df["Duplicate reason"] = leads_df.index.map(lambda i: dup_reasons.get(i, ""))
    annotated_df = leads_df

    # clean_df/suppressed_df are deliberately NOT built here: they're just
    # `annotated_df` filtered by `mask`, and materializing both up front means
    # 3 full-size copies of the list alive at once (annotated + clean +
    # suppressed) before a single output file has even been written. The
    # caller derives each one on demand, right before serializing it, and
    # drops it immediately after — see main.py's email_scrubber_endpoint.
    return {
        "total_leads": len(leads_df),
        "suppressed": int(mask.sum()),
        "clean": int((~mask).sum()),
        "duplicates": len(duplicates_df),
        "email_column_used": col,
        "annotated_df": annotated_df,
        "duplicates_df": duplicates_df,
        "suppression_mask": mask,
    }


# ---------------------------------------------------------------------------
# Tool 2: TAL : SUPP Scrubber (account-level)
# ---------------------------------------------------------------------------

def build_domain_suppression_set(content: bytes, filename: str) -> set:
    tables = read_tables(content, filename)
    domains = set()
    for df in tables.values():
        kind, col = detect_domain_column(df)
        if not col:
            if df.shape[1] == 1:
                col = df.columns[0]
                kind = "domain"
            else:
                continue
        for v in df[col].dropna().astype(str):
            if kind == "email":
                d = domain_of_email(v.strip().lower())
            else:
                d = domain_of_url(v)
            if d:
                domains.add(d)
    return domains


def tal_supp_scrubber(supp_content, supp_name, tal_content, tal_name):
    supp_domains = build_domain_suppression_set(supp_content, supp_name)
    tal_df = read_single_table(tal_content, tal_name)
    kind, col = detect_domain_column(tal_df)
    if not col:
        raise ValueError("Could not detect a domain/website/email column in the TAL file.")

    def row_domain(v):
        return domain_of_email(str(v).strip().lower()) if kind == "email" else domain_of_url(v)

    mask = tal_df[col].astype(str).apply(lambda v: row_domain(v) in supp_domains)
    clean_df = tal_df[~mask]
    suppressed_df = tal_df[mask]

    return {
        "total_tal": len(tal_df),
        "suppressed": int(mask.sum()),
        "clean": len(clean_df),
        "domain_column_used": col,
        "clean_df": clean_df,
        "suppressed_df": suppressed_df,
    }


# ---------------------------------------------------------------------------
# Tool 3: TAL : SUPP Overlap Analysis (enriched, two suppression sources)
# ---------------------------------------------------------------------------

def build_domains_from_email_supp(content: bytes, filename: str) -> set:
    tables = read_tables(content, filename)
    domains = set()
    for df in tables.values():
        if df.shape[1] == 1:
            values = df.iloc[:, 0].dropna().astype(str).tolist()
        else:
            col = detect_email_column(df)
            values = df[col].dropna().astype(str).tolist() if col else []
        for v in values:
            kind_norm = classify(v)
            if kind_norm and kind_norm[0] == "email":
                d = domain_of_email(kind_norm[1])
                if d:
                    domains.add(d)
    return domains


def tal_overlap_analysis(tal_content, tal_name, acct_supp_content, acct_supp_name,
                          email_supp_content, email_supp_name):
    tal_df = read_single_table(tal_content, tal_name).copy()
    kind, col = detect_domain_column(tal_df)
    if not col:
        raise ValueError("Could not detect a domain/website/email column in the TAL file.")

    def row_domain(v):
        return domain_of_email(str(v).strip().lower()) if kind == "email" else domain_of_url(v)

    acct_domains = build_domain_suppression_set(acct_supp_content, acct_supp_name)
    email_domains = build_domains_from_email_supp(email_supp_content, email_supp_name)

    domains_series = tal_df[col].astype(str).apply(row_domain)
    tal_df["On Account Suppression"] = domains_series.apply(lambda d: d in acct_domains)
    tal_df["On Email Suppression"] = domains_series.apply(lambda d: d in email_domains)

    return {
        "total_tal": len(tal_df),
        "on_account_supp": int(tal_df["On Account Suppression"].sum()),
        "on_email_supp": int(tal_df["On Email Suppression"].sum()),
        "domain_column_used": col,
        "enriched_df": tal_df,
    }


# ---------------------------------------------------------------------------
# Tool 4a: List Combine
# ---------------------------------------------------------------------------

def combine_lists(files):
    """files: list of (content, filename). Auto-detect email-only sources vs account sources."""
    account_frames = []
    email_values = []
    sources = 0

    for content, filename in files:
        tables = read_tables(content, filename)
        for sheet_name, df in tables.items():
            if df.empty:
                continue
            sources += 1
            col = None
            for c in df.columns:
                if "email" in str(c).lower():
                    col = c
                    break
            if col is None and df.shape[1] == 1:
                sample = df.iloc[:, 0].dropna().astype(str).head(50)
                if len(sample) and sum(1 for v in sample if EMAIL_RE.match(v.strip())) / len(sample) > 0.5:
                    col = df.columns[0]

            if col is not None:
                vals = df[col].dropna().astype(str).tolist()
                email_values.extend([v.strip().lower() for v in vals if EMAIL_RE.match(v.strip())])
            else:
                df = df.copy()
                df["__source_file__"] = f"{filename}:{sheet_name}"
                account_frames.append(df)

    email_df = pd.DataFrame({"email": sorted(set(email_values))}) if email_values else pd.DataFrame(columns=["email"])

    if account_frames:
        combined = pd.concat(account_frames, ignore_index=True, sort=False)
        dedupe_cols = []
        for c in combined.columns:
            if "company" in str(c).lower() and "name" in str(c).lower():
                dedupe_cols.append(c)
        for c in combined.columns:
            if "domain" in str(c).lower() or "website" in str(c).lower():
                dedupe_cols.append(c)
        if dedupe_cols:
            combined = combined.drop_duplicates(subset=dedupe_cols, keep="first")
        else:
            combined = combined.drop_duplicates(keep="first")
        account_df = combined
    else:
        account_df = pd.DataFrame()

    return {
        "sources": sources,
        "account_rows": len(account_df),
        "email_rows": len(email_df),
        "account_df": account_df,
        "email_df": email_df,
    }


# ---------------------------------------------------------------------------
# Tool 4b: List Split
# ---------------------------------------------------------------------------

def split_list(content: bytes, filename: str, column: str):
    df = read_single_table(content, filename)
    if column not in df.columns:
        raise ValueError(f"Column '{column}' not found. Available: {list(df.columns)}")

    groups = df.groupby(df[column].astype(str))
    zip_buffer = io.BytesIO()
    group_info = []
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, gdf in groups:
            safe_name = re.sub(r"[^A-Za-z0-9_\-]+", "_", str(name))[:80] or "group"
            csv_bytes = gdf.to_csv(index=False).encode("utf-8")
            zf.writestr(f"{safe_name}.csv", csv_bytes)
            group_info.append({"group": name, "rows": len(gdf)})
    zip_buffer.seek(0)
    return {"groups": group_info, "zip_bytes": zip_buffer.getvalue()}


# ---------------------------------------------------------------------------
# Tool 5: List Comparison (2-3 lists, venn-style overlap)
# ---------------------------------------------------------------------------

def _extract_values(content: bytes, filename: str, dedupe: bool) -> list:
    df = read_single_table(content, filename)
    if df.shape[1] == 1:
        vals = df.iloc[:, 0].dropna().astype(str).tolist()
    else:
        col = detect_email_column(df)
        vals = df[col].dropna().astype(str).tolist() if col else df.iloc[:, 0].dropna().astype(str).tolist()
    vals = [v.strip().lower() for v in vals if v.strip()]
    if dedupe:
        vals = list(dict.fromkeys(vals))
    return vals


def compare_lists(files, dedupe_within=True):
    """files: list of (content, filename), 2 or 3 entries."""
    lists = [_extract_values(c, n, dedupe_within) for c, n in files]
    membership = defaultdict(set)
    for idx, vals in enumerate(lists):
        for v in vals:
            membership[v].add(idx)

    n_lists = len(lists)
    in_all = sum(1 for s in membership.values() if len(s) == n_lists)
    in_exactly_2 = sum(1 for s in membership.values() if len(s) == 2)
    in_exactly_1 = sum(1 for s in membership.values() if len(s) == 1)
    total_unique = len(membership)

    per_list_counts = [len(v) for v in lists]

    rows = []
    for value, idx_set in membership.items():
        row = {"value": value}
        for i in range(n_lists):
            row[f"list_{i+1}"] = i in idx_set
        row["appears_in_count"] = len(idx_set)
        rows.append(row)
    result_df = pd.DataFrame(rows).sort_values("appears_in_count", ascending=False)

    return {
        "total_unique": total_unique,
        "in_all": in_all,
        "in_exactly_2": in_exactly_2,
        "in_exactly_1": in_exactly_1,
        "per_list_counts": per_list_counts,
        "result_df": result_df,
    }


# ---------------------------------------------------------------------------
# Excel export helper
# ---------------------------------------------------------------------------

def df_to_excel_bytes(sheets: dict) -> bytes:
    buf = io.BytesIO()
    # constant_memory streams rows straight to disk instead of buffering the
    # whole workbook as Python objects, which matters for large uploads.
    with pd.ExcelWriter(buf, engine="xlsxwriter", engine_kwargs={"options": {"constant_memory": True}}) as writer:
        for sheet_name, df in sheets.items():
            safe_name = sheet_name[:31]
            (df if df is not None else pd.DataFrame()).to_excel(writer, sheet_name=safe_name, index=False)
    buf.seek(0)
    return buf.getvalue()


def df_to_records(df: pd.DataFrame, limit: int = 200) -> list:
    """JSON-safe preview of a DataFrame for rendering directly in the UI
    (NaN/NaT -> None, everything stringified where needed)."""
    if df is None or df.empty:
        return []
    preview = df.head(limit).copy()
    preview = preview.astype(object).where(pd.notnull(preview), None)
    return preview.to_dict(orient="records")
